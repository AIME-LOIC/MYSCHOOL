from fastapi import FastAPI, Request, Form, UploadFile, File, Depends, Query, HTTPException, Response
from fastapi.responses import RedirectResponse, JSONResponse, FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from typing import List, Dict
import pandas as pd
import threading, time, requests, os, signal, hashlib, hmac
from datetime import datetime
import re
from openpyxl.styles import Font, PatternFill

from config import get_db, engine, Base, init_db
from model import Student, Visit, Product
from system_admin import School, create_school_database, list_all_schools, delete_school_database, get_school_database_info
from fastapi.middleware.cors import CORSMiddleware

# Initialize database tables on startup (creates all tables if they don't exist)
init_db()

app = FastAPI(title="School Visit Management System")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # or your frontend URL
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Templates & static files
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

# ==================== AUTH HELPERS ====================

SESSION_COOKIE = "admin_session"

def _make_token(school_name: str) -> str:
    secret = os.getenv("SECRET_KEY", "myschool_secret_key")
    return hmac.new(secret.encode(), school_name.encode(), hashlib.sha256).hexdigest()

def _check_admin(request: Request, school_name: str) -> bool:
    token = request.cookies.get(SESSION_COOKIE)
    return token == _make_token(school_name)


# In-memory school cache to avoid repeated DB lookups
_school_cache: dict = {}

def _get_school(school_name: str, db: Session):
    if school_name not in _school_cache:
        school = db.query(School).filter(School.school_name == school_name).first()
        if school:
            _school_cache[school_name] = {"id": school.id, "school_code": school.school_code, "school_name": school.school_name}
    return _school_cache.get(school_name)


# ==================== SYSTEM ADMIN ENDPOINTS ====================

# System Admin - Main dashboard
@app.get("/myadmin")
def system_admin_dashboard(request: Request):
    """System admin main portal for managing schools"""
    schools = list_all_schools()
    return templates.TemplateResponse("system_admin.html", {
        "request": request,
        "schools": schools,
        "total_schools": len(schools)
    })


# System Admin - Create new school
@app.post("/myadmin/create-school")
def create_new_school(
    school_name: str = Form(...),
    school_code: str = Form(...),
    db: Session = Depends(get_db)
):
    """Create a new school and its associated database"""
    # Check if school already exists
    existing = db.query(School).filter(School.school_name == school_name).first()
    if existing:
        return {"status": "error", "message": "School already exists"}
    
    # Create school in system database
    school = School(school_name=school_name, school_code=school_code)
    db.add(school)
    db.commit()
    
    # Create school-specific database
    result = create_school_database(school_name, school_code)
    
    if result["status"] == "success":
        return {
            "status": "success",
            "message": f"School '{school_name}' created successfully",
            "school_id": school.id,
            "school_name": school_name,
            "school_code": school_code
        }
    else:
        # Rollback if database creation fails
        db.delete(school)
        db.commit()
        return result


# System Admin - List all schools
@app.get("/myadmin/schools")
def get_all_schools(db: Session = Depends(get_db)):
    """Get list of all registered schools"""
    schools = db.query(School).filter(School.is_active == 1).all()
    return {
        "status": "success",
        "total": len(schools),
        "schools": [
            {
                "id": s.id,
                "school_name": s.school_name,
                "school_code": s.school_code,
                "created_at": s.created_at,
                "is_active": s.is_active
            } for s in schools
        ]
    }


# System Admin - Delete school
@app.delete("/myadmin/schools/{school_id}")
def delete_school(school_id: int, db: Session = Depends(get_db)):
    """Delete a school and its data"""
    school = db.query(School).filter(School.id == school_id).first()
    if not school:
        return {"status": "error", "message": "School not found"}
    
    # Delete from database
    db.delete(school)
    db.commit()
    
    # Delete school-specific database
    result = delete_school_database(school.school_name)
    
    return {
        "status": "success",
        "message": f"School '{school.school_name}' deleted successfully"
    }


# ==================== SCHOOL-SPECIFIC ENDPOINTS ====================

# Parent chooses type for specific school
@app.get("/{school_name}/parentportal")
def parent_choice(school_name: str, request: Request, db: Session = Depends(get_db)):
    """Parent portal for a specific school"""
    school = db.query(School).filter(School.school_name == school_name).first()
    if not school:
        raise HTTPException(status_code=404, detail="School not found")
    
    return templates.TemplateResponse("index.html", {
        "request": request,
        "school_name": school_name,
        "school_code": school.school_code
    })


# Admin login page
@app.get("/{school_name}/admin/login")
def admin_login_page(school_name: str, request: Request, db: Session = Depends(get_db)):
    school = db.query(School).filter(School.school_name == school_name).first()
    if not school:
        raise HTTPException(status_code=404, detail="School not found")
    if _check_admin(request, school_name):
        return RedirectResponse(f"/{school_name}/admin", status_code=302)
    return templates.TemplateResponse("login.html", {"request": request, "school_name": school_name, "error": ""})


@app.post("/{school_name}/admin/login")
def admin_login(
    school_name: str,
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(get_db)
):
    school = db.query(School).filter(School.school_name == school_name).first()
    if not school:
        raise HTTPException(status_code=404, detail="School not found")
    expected_user = school_name
    expected_pass = f"{school_name}001"
    if username == expected_user and password == expected_pass:
        token = _make_token(school_name)
        response = RedirectResponse(f"/{school_name}/admin", status_code=302)
        response.set_cookie(SESSION_COOKIE, token, httponly=True, samesite="lax", max_age=86400 * 7)
        return response
    return templates.TemplateResponse("login.html", {"request": request, "school_name": school_name, "error": "Invalid username or password"})


@app.get("/{school_name}/admin/logout")
def admin_logout(school_name: str):
    response = RedirectResponse(f"/{school_name}/admin/login", status_code=302)
    response.delete_cookie(SESSION_COOKIE)
    return response


# Admin dashboard for specific school
@app.get("/{school_name}/admin")
def admin_choice(school_name: str, request: Request, db: Session = Depends(get_db)):
    """Admin dashboard for a specific school"""
    school = db.query(School).filter(School.school_name == school_name).first()
    if not school:
        raise HTTPException(status_code=404, detail="School not found")
    if not _check_admin(request, school_name):
        return RedirectResponse(f"/{school_name}/admin/login", status_code=302)
    return templates.TemplateResponse("admin.html", {
        "request": request,
        "school_name": school_name,
        "school_code": school.school_code
    })


# ==================== SCHOOL-SPECIFIC API ENDPOINTS ====================

@app.get("/{school_name}/students/search", response_model=List[Dict])
def search_students(school_name: str, q: str = Query(..., min_length=1), db: Session = Depends(get_db)):
    school = _get_school(school_name, db)
    if not school:
        raise HTTPException(status_code=404, detail="School not found")
    query = db.query(Student.id, Student.student_name, Student.class_name).filter(
        (Student.school_id == school["id"]) | (Student.school_id.is_(None)),
        Student.student_name.ilike(f"%{q}%")
    ).limit(20).all()
    return [{"id": s.id, "student_name": s.student_name, "class_name": s.class_name} for s in query]


@app.post("/{school_name}/visits/add")
def add_visit(
    school_name: str,
    student_id: int = Form(...),
    visit_type: str = Form(...),
    movement_method: str | None = Form(None),
    plate_number: str | None = Form(None),
    db: Session = Depends(get_db)
):
    """Add a visit for a student in a specific school"""
    school = db.query(School).filter(School.school_name == school_name).first()
    if not school:
        raise HTTPException(status_code=404, detail="School not found")
    
    try:
        student = db.query(Student).filter(Student.id == student_id).first()
        if not student:
            return {"status": "error", "message": "Student not found"}
        if student.school_id is not None and student.school_id != school.id:
            return {"status": "error", "message": "Student does not belong to this school"}

        if visit_type not in {"visit_day", "parent_meeting"}:
            return {"status": "error", "message": "Invalid visit type"}

        is_sos_school = (
            (school.school_code or "").strip().lower() == "sos"
            or "sos" in (school.school_name or "").strip().lower()
        )

        movement_method_clean = None
        plate_number_clean = None
        if is_sos_school:
            movement_method_clean = (movement_method or "").strip().lower()
            if movement_method_clean not in {"with_car", "without_car"}:
                return {"status": "error", "message": "Please select movement method"}

            if movement_method_clean == "with_car":
                plate_number_clean = (plate_number or "").strip().upper()
                if not plate_number_clean:
                    return {"status": "error", "message": "Plate number is required"}

                # Normalize by stripping spaces and validate format: RA + letter + 3 digits + letter (e.g. REA 123A)
                plate_number_clean = re.sub(r"\s+", "", plate_number_clean)
                if not re.match(r"^RA[A-Z]\d{3}[A-Z]$", plate_number_clean):
                    return {"status": "error", "message": "Plate number must be in the format RAx 123A (start with RA, then a letter, 3 digits, and a letter)."}
            else:
                plate_number_clean = None

        today = datetime.now().date()

        existing = db.query(Visit).filter(
            Visit.student_id == student.id,
            Visit.visit_type == visit_type,
            Visit.visit_date == today
        ).first()
        if existing:
            return {"status": "error", "message": "Already recorded today"}

        visit = Visit(
            student_id=student.id,
            visit_type=visit_type,
            visit_date=today,
            status="done",
            movement_method=movement_method_clean,
            arrival_plate_number=plate_number_clean,
        )
        db.add(visit)
        db.commit()
        return {"status": "success", "visit_id": visit.id}
    except Exception as e:
        db.rollback()
        return {"status": "error", "message": str(e)}


@app.get("/{school_name}/admin/data/{visit_type}")
def admin_data(school_name: str, visit_type: str, visit_date: str | None = Query(None), db: Session = Depends(get_db)):
    school = _get_school(school_name, db)
    if not school:
        raise HTTPException(status_code=404, detail="School not found")

    query = db.query(
        Visit.id, Visit.visit_date, Visit.created_at,
        Visit.movement_method, Visit.arrival_plate_number, Visit.assigned_plate_number,
        Student.student_name, Student.class_name
    ).join(Student).filter(
        Visit.visit_type == visit_type,
        (Student.school_id == school["id"]) | (Student.school_id.is_(None))
    )

    if visit_date:
        try:
            target_date = datetime.strptime(visit_date, "%Y-%m-%d").date()
            query = query.filter(Visit.visit_date == target_date)
        except ValueError:
            pass

    rows = query.order_by(Visit.visit_date.desc(), Visit.id.desc()).all()

    result = {}
    for r in rows:
        cls = r.class_name
        if cls not in result:
            result[cls] = []
        result[cls].append({
            "student_name": r.student_name,
            "date": r.visit_date.strftime("%Y-%m-%d"),
            "timestamp": r.created_at.isoformat() if r.created_at else None,
            "id": r.id,
            "movement_method": r.movement_method,
            "arrival_plate_number": r.arrival_plate_number,
            "assigned_plate_number": r.assigned_plate_number,
        })

    stats = {cls: len(students) for cls, students in result.items()}
    return {"data": result, "stats": stats, "total": len(rows)}


@app.post("/{school_name}/admin/upload-students")
def upload_students(school_name: str, file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Upload students for a specific school from Excel file"""
    school = db.query(School).filter(School.school_name == school_name).first()
    if not school:
        raise HTTPException(status_code=404, detail="School not found")
    
    try:
        df = pd.read_excel(file.file)
    except Exception:
        return JSONResponse({"error": "Invalid Excel file"}, status_code=400)

    required_columns = {"student_name", "class_name"}
    if not required_columns.issubset(df.columns):
        return JSONResponse({"error": "Invalid Excel format"}, status_code=400)

    added = 0
    skipped = 0
    for _, row in df.iterrows():
        name, cls = row["student_name"], row["class_name"]
        if db.query(Student).filter(Student.student_name == name, Student.class_name == cls, Student.school_id == school.id).first():
            skipped += 1
            continue
        db.add(Student(student_name=name, class_name=cls, school_id=school.id))
        added += 1
    db.commit()
    return {"status": "success", "added": added, "skipped": skipped}


# Get all students for a specific school
@app.get("/{school_name}/admin/students")
def get_all_students(school_name: str, db: Session = Depends(get_db)):
    school = _get_school(school_name, db)
    if not school:
        raise HTTPException(status_code=404, detail="School not found")
    students = db.query(Student.id, Student.student_name, Student.class_name).filter(
        (Student.school_id == school["id"]) | (Student.school_id.is_(None))
    ).all()
    return {"status": "success", "students": [{"id": s.id, "student_name": s.student_name, "class_name": s.class_name} for s in students]}


# Delete a student by ID from a specific school
@app.delete("/{school_name}/admin/students/{student_id}")
def delete_student(school_name: str, student_id: int, db: Session = Depends(get_db)):
    """Delete a student from a specific school"""
    school = db.query(School).filter(School.school_name == school_name).first()
    if not school:
        raise HTTPException(status_code=404, detail="School not found")
    
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        return {"status": "error", "message": "Student not found"}

    db.delete(student)
    db.commit()
    return {"status": "success", "message": f"Student '{student.student_name}' deleted"}


# Delete a visit by ID
@app.delete("/{school_name}/admin/visits/{visit_id}")
def delete_visit(school_name: str, visit_id: int, db: Session = Depends(get_db)):
    school = db.query(School).filter(School.school_name == school_name).first()
    if not school:
        raise HTTPException(status_code=404, detail="School not found")
    visit = db.query(Visit).join(Student).filter(
        Visit.id == visit_id,
        (Student.school_id == school.id) | (Student.school_id.is_(None))
    ).first()
    if not visit:
        return {"status": "error", "message": "Visit not found"}
    db.delete(visit)
    db.commit()
    return {"status": "success"}


# Add a single student to a specific school
@app.post("/{school_name}/admin/add_student")
def add_student(
    school_name: str,
    student_name: str = Form(...),
    class_name: str = Form(...),
    db: Session = Depends(get_db)
):
    """Add a single student to a specific school"""
    school = db.query(School).filter(School.school_name == school_name).first()
    if not school:
        raise HTTPException(status_code=404, detail="School not found")
    
    # Check if student already exists
    existing = db.query(Student).filter(
        Student.student_name == student_name,
        Student.class_name == class_name,
        Student.school_id == school.id
    ).first()
    if existing:
        return {"status": "error", "message": "Student already exists"}

    # Add new student with school_id
    student = Student(student_name=student_name, class_name=class_name, school_id=school.id)
    db.add(student)
    db.commit()
    return {"status": "success", "student_id": student.id, "message": "Student added successfully"}


@app.get("/{school_name}/admin/car-management")
def get_car_management_data(
    school_name: str,
    visit_type: str = Query(..., pattern="^(visit_day|parent_meeting)$"),
    visit_date: str | None = Query(None),
    db: Session = Depends(get_db)
):
    """List SOS visits (default today) with car movement details for admin assignment."""
    school = db.query(School).filter(School.school_name == school_name).first()
    if not school:
        raise HTTPException(status_code=404, detail="School not found")
    is_sos_school = (
        (school.school_code or "").strip().lower() == "sos"
        or "sos" in (school.school_name or "").strip().lower()
    )
    if not is_sos_school:
        return {"status": "error", "message": "Car management is available only for SOS school"}

    target_date = datetime.now().date()
    if visit_date:
        try:
            target_date = datetime.strptime(visit_date, "%Y-%m-%d").date()
        except ValueError:
            return {"status": "error", "message": "Invalid date format. Use YYYY-MM-DD"}

    visits = db.query(Visit).join(Student).filter(
        Visit.visit_type == visit_type,
        Visit.visit_date == target_date,
        (Student.school_id == school.id) | (Student.school_id.is_(None))
    ).order_by(Visit.id.asc()).all()

    return {
        "status": "success",
        "records": [
            {
                "visit_id": v.id,
                "student_id": v.student.id,
                "student_name": v.student.student_name,
                "class_name": v.student.class_name,
                "visit_type": v.visit_type,
                "visit_date": v.visit_date.strftime("%Y-%m-%d"),
                "movement_method": v.movement_method,
                "arrival_plate_number": v.arrival_plate_number,
                "assigned_plate_number": v.assigned_plate_number,
            }
            for v in visits
        ],
    }


@app.post("/{school_name}/admin/car-management/assign")
def assign_car_plate(
    school_name: str,
    visit_id: int = Form(...),
    assigned_plate_number: str = Form(...),
    db: Session = Depends(get_db)
):
    """Assign a managed car plate to a specific student visit."""
    school = db.query(School).filter(School.school_name == school_name).first()
    if not school:
        raise HTTPException(status_code=404, detail="School not found")
    is_sos_school = (
        (school.school_code or "").strip().lower() == "sos"
        or "sos" in (school.school_name or "").strip().lower()
    )
    if not is_sos_school:
        return {"status": "error", "message": "Car management is available only for SOS school"}

    visit = db.query(Visit).join(Student).filter(
        Visit.id == visit_id,
        (Student.school_id == school.id) | (Student.school_id.is_(None))
    ).first()
    if not visit:
        return {"status": "error", "message": "Visit record not found"}

    plate = assigned_plate_number.strip().upper()
    if not plate:
        return {"status": "error", "message": "Assigned plate number is required"}

    visit.assigned_plate_number = plate
    db.commit()
    return {"status": "success", "message": "Car plate assigned successfully", "visit_id": visit.id}


# ==================== PRODUCT ENDPOINTS ====================

@app.get("/{school_name}/products")
def get_products(school_name: str, db: Session = Depends(get_db)):
    school = db.query(School).filter(School.school_name == school_name).first()
    if not school:
        raise HTTPException(status_code=404, detail="School not found")
    products = db.query(Product).filter(
        (Product.school_id == school.id) | (Product.school_id.is_(None))
    ).all()
    return {"status": "success", "products": [{"product_id": p.product_id, "product_name": p.product_name, "product_price": p.product_price} for p in products]}


@app.post("/{school_name}/admin/products/add")
def add_product(
    school_name: str,
    product_name: str = Form(...),
    product_price: int = Form(...),
    db: Session = Depends(get_db)
):
    school = db.query(School).filter(School.school_name == school_name).first()
    if not school:
        raise HTTPException(status_code=404, detail="School not found")
    product = Product(product_name=product_name, product_price=product_price, school_id=school.id)
    db.add(product)
    db.commit()
    return {"status": "success", "product_id": product.product_id}


@app.put("/{school_name}/admin/products/{product_id}")
def update_product(
    school_name: str,
    product_id: int,
    product_name: str = Form(...),
    product_price: int = Form(...),
    db: Session = Depends(get_db)
):
    school = db.query(School).filter(School.school_name == school_name).first()
    if not school:
        raise HTTPException(status_code=404, detail="School not found")
    product = db.query(Product).filter(Product.product_id == product_id, Product.school_id == school.id).first()
    if not product:
        return {"status": "error", "message": "Product not found"}
    product.product_name = product_name
    product.product_price = product_price
    db.commit()
    return {"status": "success"}


@app.delete("/{school_name}/admin/products/{product_id}")
def delete_product(
    school_name: str,
    product_id: int,
    db: Session = Depends(get_db)
):
    school = db.query(School).filter(School.school_name == school_name).first()
    if not school:
        raise HTTPException(status_code=404, detail="School not found")
    product = db.query(Product).filter(Product.product_id == product_id, Product.school_id == school.id).first()
    if not product:
        return {"status": "error", "message": "Product not found"}
    db.delete(product)
    db.commit()
    return {"status": "success"}


# ==================== EXPORT ENDPOINTS ====================

@app.get("/{school_name}/export/students")
def export_students_to_excel(
    school_name: str,
    class_name: str | None = Query(None),
    db: Session = Depends(get_db)
):
    """Export all students for a specific school to Excel file"""
    school = db.query(School).filter(School.school_name == school_name).first()
    if not school:
        raise HTTPException(status_code=404, detail="School not found")
    
    try:
        query = db.query(Student).filter(
            (Student.school_id == school.id) | (Student.school_id.is_(None))
        )
        class_filter = (class_name or "").strip()
        if class_filter and class_filter.lower() != "all":
            query = query.filter(Student.class_name == class_filter)
        students = query.all()
        
        # Prepare data for Excel
        data = []
        for student in students:
            data.append({
                "Student Name": student.student_name,
                "Class": student.class_name,
                "ID": student.id
            })
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Generate filename with timestamp
        filename = f"students_{school_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = f"/tmp/{filename}"
        
        # Write to Excel with formatting
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Students', index=False)
            
            # Access the workbook and worksheet for formatting
            workbook = writer.book
            worksheet = writer.sheets['Students']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Style header row
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
        
        return FileResponse(filepath, filename=filename, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/{school_name}/export/visits")
def export_visits_to_excel(
    school_name: str, 
    visit_type: str = Query(..., pattern="^(visit_day|parent_meeting)$"),
    visit_date: str | None = Query(None),
    db: Session = Depends(get_db)
):
    """Export visit data for a specific school and visit type to Excel file"""
    school = db.query(School).filter(School.school_name == school_name).first()
    if not school:
        raise HTTPException(status_code=404, detail="School not found")
    
    try:
        query = db.query(Visit).join(Student).filter(
            Visit.visit_type == visit_type,
            (Student.school_id == school.id) | (Student.school_id.is_(None))
        )
        if visit_date:
            try:
                target = datetime.strptime(visit_date, "%Y-%m-%d").date()
                query = query.filter(Visit.visit_date == target)
            except ValueError:
                pass
        visits = query.all()
        
        # Prepare data for Excel
        data = []
        for visit in visits:
            data.append({
                "Student Name": visit.student.student_name,
                "Class": visit.student.class_name,
                "Visit Type": visit.visit_type.replace('_', ' ').title(),
                "Visit Date": visit.visit_date.strftime("%Y-%m-%d"),
                "Status": visit.status,
                "Movement Method": visit.movement_method if visit.movement_method else "-",
                "Arrival Plate": visit.arrival_plate_number if visit.arrival_plate_number else "-",
                "Assigned Plate": visit.assigned_plate_number if visit.assigned_plate_number else "-",
            })
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Generate filename with timestamp
        visit_type_label = "visit_day" if visit_type == "visit_day" else "parent_meeting"
        filename = f"{visit_type_label}_{school_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = f"/tmp/{filename}"
        
        # Write to Excel with formatting
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Visit Data', index=False)
            
            # Access the workbook and worksheet for formatting
            workbook = writer.book
            worksheet = writer.sheets['Visit Data']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Style header row
            header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
        
        return FileResponse(filepath, filename=filename, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


import threading
import time
import requests
import os

def keep_awake():
    url = os.environ.get("RENDER_URL", "https://myschool-rw-web.onrender.com")  # set your deployed URL in env variables
    while True:
        try:
            requests.get(url)
        except:
            pass
        time.sleep(10 * 60)  # ping every 10 minutes

threading.Thread(target=keep_awake, daemon=True).start()
  
